import time
import openpyxl
import pandas as pd

import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.proyectos_utils import find_row_contains, extraer_actividades

if __name__ == "__main__":
    start = time.time()
    ruta_excel = r"C:\Users\Adminstrador\Documents\3 - Ave Fenix\Propuesta\Seguimiento a Procesos y Proyectos de Calidad Educativa.xlsx"

    hojas = [
        "Juegos_Pro_2024", "Gestion_Escolar_Pro_2024", "Convivencia_Pro_2024",
        "Educacion_Inicial_Pro_2024", "Articulacion_Pro_2024", "Normales_Pro_2024",
        "PEER_Pro_2024", "PILEO_Pro_2024", "Familia_Escuela_Pro_2024",
        "SEIP_Pro_2024", "AFRO_Pro_2024", "ADULTOS_Pro_2024",
        "PADRINAZGO_Pro_2024", "BIBLIOTECA_Pro_2024"
    ]

    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    proyectos, actividades_total = [], []

    for hoja in hojas:
        if hoja not in wb.sheetnames:
            print(f"❌ La hoja {hoja} no existe en el archivo.")
            continue

        ws = wb[hoja]
        print(f"\n📄 Procesando hoja: {hoja}")

        info_proyecto = {
            "Nombre del Proyecto": ws["D3"].value,
            "Vigencia": ws["D4"].value,
            "Código BPIN": ws["F4"].value,
            "Código PI": ws["H4"].value,
            "Total Ejecutado": ws["J4"].value,
            "RECURSOS": ws["L4"].value,
            "Hoja": hoja
        }

        r_title = find_row_contains(ws, r"ACTIVIDADES", 1, 200)
        header_row = (r_title or 0) + 1
        data_start = header_row + 1
        r_obs = find_row_contains(ws, r"Observaciones\s+Generales", start_row=data_start, end_row=500)

        col_idx = [2, 3, 6, 7, 8, 9, 10, 11, 12]
        columnas = [
            "N°", "Actividad del Proyecto", "Total Ejecutado 2024",
            "Componente PAM", "¿A qué actor va dirigida?",
            "Número de Beneficiarios", "Entrega Dotación (SI / NO)",
            "Descripción de la Dotación Entregada", "Evidencia de la Actividad"
        ]

        df_actividades = extraer_actividades(ws, hoja, col_idx, columnas, r_title, r_obs, data_start)
        df_actividades["Nombre del Proyecto"] = info_proyecto["Nombre del Proyecto"]
        df_actividades["Código PI"] = info_proyecto["Código PI"]

        proyectos.append(info_proyecto)
        actividades_total.append(df_actividades)

    df_proyectos = pd.DataFrame(proyectos)
    df_actividades = pd.concat(actividades_total, ignore_index=True) if actividades_total else pd.DataFrame()

    output_file = "outputs/Proyectos_Actividades_POAI_2024.xlsx"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_proyectos.to_excel(writer, sheet_name="Info_Proyectos", index=False)
        df_actividades.to_excel(writer, sheet_name="Actividades", index=False)

    end = time.time()
    print(f"\n✅ POAI 2024: {len(df_proyectos)} proyectos y {len(df_actividades)} actividades en {end - start:.2f} seg → {output_file}")
