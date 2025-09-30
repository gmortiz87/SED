import openpyxl
import pandas as pd
from pathlib import Path
import re

# === Configuración de columnas por fuente ===
COLUMN_MAPS = {
    "POAI": {
        "idx": [2, 3, 6, 7, 8, 9, 10, 11, 12],
        "cols": [
            "N°", "Actividad del Proyecto", "Total Ejecutado",
            "Componente PAM", "¿A qué actor va dirigida?",
            "Número de Beneficiarios", "Entrega Dotación (SI / NO)",
            "Descripción de la Dotación Entregada", "Evidencia de la Actividad"
        ]
    },
    "REGALIAS": {
        "idx": [2, 3, 6, 7, 8, 9, 10, 11, 12],
        "cols": [
            "N°", "Actividad del Proyecto", "Total Ejecutado",
            "Componente PAM", "¿A qué actor va dirigida?",
            "Número de Beneficiarios", "Entrega Dotación (SI / NO)",
            "Descripción de la Dotación Entregada", "Evidencia de la Actividad"
        ]
    },
    "ESTRATEGIAS": {
        "idx": [2, 3, 6, 7, 8, 9, 12],
        "cols": [
            "N°", "Actividad del Proyecto", "¿A qué actor va dirigida?",
            "Número de Beneficiarios", "Entrega Dotación (SI / NO)",
            "Descripción de la Dotación Entregada", "Evidencia de la Actividad"
        ]
    }
}


def find_row_contains(ws, pattern, start_row=1, end_row=200):
    regex = re.compile(pattern, re.IGNORECASE)
    for row in range(start_row, end_row + 1):
        for cell in ws[row]:
            if cell.value and regex.search(str(cell.value)):
                return row
    return None


def procesar_fuente_proyectos(ruta_excel, fuente, hojas, output_file):
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    proyectos, actividades_total, metas_total = [], [], []   # todas las listas aquí

    for hoja in hojas:
        if hoja not in wb.sheetnames:
            print(f"[ALERTA] La hoja {hoja} no existe en el archivo.")
            continue

        ws = wb[hoja]

        # === Metadatos del proyecto ===
        if fuente.upper().startswith("POAI"):
            info_proyecto = {
                "Nombre_Proyecto": ws["D3"].value,
                "Vigencia": ws["D4"].value,
                "Código BPIN": ws["F4"].value,
                "Código PI": ws["H4"].value,
                "Total Ejecutado": ws["J4"].value,
                "RECURSOS": ws["L4"].value,
                "Hoja": hoja
            }
        elif fuente.upper() == "REGALIAS":
            info_proyecto = {
                "Nombre_Proyecto": ws["D3"].value,
                "Vigencia": ws["D4"].value,
                "Código BPIN": ws["F4"].value,
                "Código PI": ws["H4"].value,
                "Total Ejecutado": ws["J4"].value,
                "RECURSOS": ws["L4"].value,
                "Hoja": hoja
            }
        elif fuente.upper() == "ESTRATEGIAS":
            info_proyecto = {
                "Nombre_Proyecto": ws["D3"].value,
                "Hoja": hoja
            }
        else:
            print(f"[ERROR] Fuente {fuente} no soportada.")
            continue

        # === Localizar secciones ===
        r_title = find_row_contains(ws, r"ACTIVIDADES", 1, 200)
        header_row = (r_title or 0) + 1
        data_start = header_row + 1
        r_obs = find_row_contains(ws, r"Observaciones\s+Generales", start_row=data_start, end_row=500)

        # === Configuración de columnas según fuente ===
        fuente_upper = fuente.upper()
        if fuente_upper.startswith("POAI"):
            config = COLUMN_MAPS["POAI"]
        elif fuente_upper == "REGALIAS":
            config = COLUMN_MAPS["REGALIAS"]
        elif fuente_upper == "ESTRATEGIAS":
            config = COLUMN_MAPS["ESTRATEGIAS"]
        else:
            print(f"[ERROR] No hay configuración de columnas para la fuente {fuente}")
            continue

        col_idx = config["idx"]
        columnas = config["cols"]

        # === Extraer metas del producto (solo si no es ESTRATEGIAS) ===
        if fuente_upper != "ESTRATEGIAS":
            nombre_proy = ws["D3"].value
            for r in range(5, 31):  # filas 4..30
                b = ws.cell(row=r, column=2).value  # Columna B
                if not b:
                    continue
                b_txt = str(b).strip()
                if re.match(r"^Met", b_txt, re.IGNORECASE):
                    desc_parts = []
                    for col in range(4, 13):  # columnas D..L
                        v = ws.cell(row=r, column=col).value
                        if v not in (None, ""):
                            desc_parts.append(str(v).strip())
                    desc = " ".join(desc_parts).strip() or "(sin descripción)"
                    metas_total.append({
                        "Hoja": hoja,
                        "Nombre_Proyecto": nombre_proy,
                        "Meta": b_txt,
                        "Descripción": desc
                    })

        # === Extraer actividades ===
        data = []
        row = data_start
        last_row = (r_obs - 1) if r_obs else ws.max_row

        while row <= last_row:
            values = [ws.cell(row=row, column=i).value for i in col_idx]

            # Extraer URL de evidencia
            evidencia_cell = ws.cell(row=row, column=col_idx[-1])
            evidencia_url = None

            # Caso 1: hipervínculo directo
            if evidencia_cell.hyperlink:
                evidencia_url = evidencia_cell.hyperlink.target

            # Caso 2: revisar hipervínculos de la hoja
            elif hasattr(ws, "_hyperlinks"):
                for hl in ws._hyperlinks:
                    if evidencia_cell.coordinate in hl.ref:
                        evidencia_url = hl.target
                        break

            # Caso 3: texto con URL
            if not evidencia_url and evidencia_cell.value and isinstance(evidencia_cell.value, str):
                if evidencia_cell.value.strip().lower().startswith(("http://", "https://")):
                    evidencia_url = evidencia_cell.value.strip()

            fila = values + [evidencia_url]
            if any(fila):
                data.append(fila)
            row += 1

        df_actividades = pd.DataFrame(data, columns=columnas + ["Evidencia_URL"])

        # === Observaciones Generales ===
        if r_obs:
            obs_parts = []
            for col in range(5, 13):
                v = ws.cell(row=r_obs, column=col).value
                if v not in (None, ""):
                    obs_parts.append(str(v).strip())
            observacion_final = " | ".join(obs_parts).strip() if obs_parts else "Sin observaciones"
        else:
            observacion_final = "Sin observaciones"

        df_actividades["Observaciones Generales"] = observacion_final

        # === Limpieza encabezados falsos ===
        if not df_actividades.empty and isinstance(df_actividades.iloc[0, 1], str) and "actividad" in df_actividades.iloc[0, 1].lower():
            df_actividades = df_actividades.iloc[1:].reset_index(drop=True)

        # === Metadatos adicionales ===
        df_actividades["Hoja"] = hoja
        df_actividades["Nombre_Proyecto"] = ws["D3"].value

        proyectos.append(info_proyecto)
        actividades_total.append(df_actividades)

    # === DataFrames finales ===
    df_proyectos = pd.DataFrame(proyectos)
    actividades_total = [df for df in actividades_total if not df.empty]
    df_actividades = pd.concat(actividades_total, ignore_index=True) if actividades_total else pd.DataFrame()
    df_metas = pd.DataFrame(metas_total)

    # === Exportar resultados ===
    if not df_proyectos.empty or not df_actividades.empty or not df_metas.empty:
        output_path = Path("extraccion/outputs") / output_file
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_proyectos.to_excel(writer, sheet_name="Info_Proyectos", index=False)
            df_actividades.to_excel(writer, sheet_name="Actividades", index=False)
            if not df_metas.empty:
                df_metas.to_excel(writer, sheet_name="Metas", index=False)

        print(f"[OK] {len(df_proyectos)} proyectos, "
              f"{len(df_actividades)} actividades y "
              f"{len(df_metas)} metas exportadas a {output_path}")

    # === Return condicional ===
    if fuente.upper() == "ESTRATEGIAS":
        return df_proyectos, df_actividades
    else:
        return df_proyectos, df_actividades, df_metas
