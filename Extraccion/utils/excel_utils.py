import openpyxl
import pandas as pd

def extraer_datos(file_path, hoja, fila_titulos=2, fila_datos=3):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if hoja not in wb.sheetnames:
        print(f"La hoja '{hoja}' no existe en el archivo.")
        return pd.DataFrame()

    ws = wb[hoja]
    print(f"[INFO] Procesando hoja: {hoja}")

    headers = [
        str(cell.value).strip() if cell.value else f"Col{i+1}"
        for i, cell in enumerate(ws[fila_titulos])
    ]

    all_rows = []
    for row in ws.iter_rows(min_row=fila_datos):
        valores = [
            cell.hyperlink.target if cell.hyperlink else cell.value
            for cell in row
        ]
        if not all(v is None or v == "" for v in valores):
            fila_dict = dict(zip(headers, valores))
            fila_dict["Hoja"] = hoja
            all_rows.append(fila_dict)

    return pd.DataFrame(all_rows)
