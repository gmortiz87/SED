import pandas as pd
from openpyxl import load_workbook
from config_fuentes import FUENTES_HOJAS_IGP
from extraccion.utils.export_utils import guardar_excel
from .helpers_igp import procesar_igp_hoja


def run():
    """
    Ejecuta la extracción de información de metas del producto (IGP)
    para la fuente REGALIAS, según las hojas configuradas en config_fuentes.
    """
    fuente = "REGALIAS"
    archivo = f"extraccion/inputs/{fuente}.xlsx"
    hojas = FUENTES_HOJAS_IGP.get(fuente, [])

    print(f"\n📘 Iniciando extracción IGP para {fuente}...")
    wb = load_workbook(archivo, data_only=True)
    registros = []

    for hoja in hojas:
        if hoja not in wb.sheetnames:
            print(f"[ALERTA] La hoja '{hoja}' no existe en el archivo {fuente}.")
            continue

        ws = wb[hoja]
        print(f"[INFO] Procesando hoja: {hoja}")
        df_hoja = procesar_igp_hoja(ws, hoja, fuente)

        if df_hoja is not None and not df_hoja.empty:
            registros.append(df_hoja)
            print(f"[OK] {hoja}: {len(df_hoja)} registros extraídos.")
        else:
            print(f"[AVISO] {hoja}: sin registros válidos.")

    # === Exportar resultados ===
    if registros:
        df_final = pd.concat(registros, ignore_index=True)
        guardar_excel({"stg_meta_producto": df_final}, f"IGP_{fuente}.xlsx")
        print(f"[✅] Extracción completada para {fuente}: {len(df_final)} registros totales.")
    else:
        print(f"[⚠️] No se generaron registros para {fuente}.")
